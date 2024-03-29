import datetime
import csv
import openpyxl
import openpyxl.worksheet
from statistics import median
from statistics import variance
from typing import List

# TODO: Implement weighted average and weighted median based on the time spent in each state
# TODO: Implement variance

START_DATE = datetime.date.fromisoformat("2024-01-01")
INPUT_FILE = "./daylio_export_2024_03_29.csv"
OUTPUT_FILE = "mood_stats.xlsx"
FULL_DATE = 0
TIME = 3
MOOD = 4
ACTIVITIES = 5


class Day:
    def __init__(self, date):
        self.date = date
        self.entries: List[Entry] = []
        self.min_value = None
        self.max_value = None
        self.avg_value = None
        self.mdn_value = None
        self.var_value = None

    def __str__(self):
        return f"{self.date} | min: {self.min_value} | max: {self.max_value} | avg: {self.avg_value} | mdn: {self.mdn_value}"

    def add_entry(self, time, mood, activities):
        # Insert at start of list so that the earliest time is at index 0
        self.entries.insert(0, Entry(self.date, time, mood, activities))

    def min(self):
        if self.min_value is not None:
            return self.min_value
        self.min_value = self.entries[0].mood
        for entry in self.entries[1:]:
            if entry.mood < self.min_value:
                self.min_value = entry.mood
        return self.min_value

    def max(self):
        if self.max_value is not None:
            return self.max_value
        self.max_value = self.entries[0].mood
        for entry in self.entries[1:]:
            if entry.mood > self.max_value:
                self.max_value = entry.mood
        return self.max_value

    def avg(self):
        if self.avg_value is not None:
            return self.avg_value
        sum = 0
        for entry in self.entries:
            sum += entry.mood
        self.avg_value = round(sum / len(self.entries), 1)
        return self.avg_value

    def mdn(self):
        if self.mdn_value is not None:
            return self.mdn_value
        mood_list = []
        for entry in self.entries:
            mood_list.append(entry.mood)
        self.mdn_value = round(median(mood_list), 1)
        return self.mdn_value

    def var(self):
        if self.var_value is not None:
            return self.var_value
        raise NotImplementedError()


class Entry:
    def __init__(self, date, time, mood, activities):
        self.date = date
        self.time = time
        self.mood = mood
        self.activities = activities

    def __str__(self):
        return f"{self.date} | {self.time} | {self.mood}"

    def activities_to_string(self):
        result = self.activities[0]
        for activity in self.activities[1:]:
            result += f" | {activity}"
        return result


days = {}
with open(INPUT_FILE) as csv_file:
    reader = csv.reader(csv_file, delimiter=";")
    next(reader)  # Skip first row
    for row in reader:
        date = datetime.date.fromisoformat(row[FULL_DATE])
        if date < START_DATE:
            break
        if date not in days.keys():
            days[date] = Day(date)
        time = datetime.time.fromisoformat(row[TIME])
        mood = int(row[MOOD].split(" ")[0])
        activities = row[ACTIVITIES].split(" | ")
        days[date].add_entry(time, mood, activities)

wb = openpyxl.Workbook()
wb.create_sheet("Entries")
wb.create_sheet("Days")
del wb["Sheet"]
days_ws = wb["Days"]
entries_ws = wb["Entries"]
headers = ["date", "min", "max", "avg", "mdn"]
days_ws.append(headers)
headers = ["date", "time", "mood", "activities"]
entries_ws.append(headers)
for day in days.values():
    row = [day.date, day.min(), day.max(), day.avg(), day.mdn()]
    days_ws.append(row)
    for entry in day.entries:
        row = [entry.date, entry.time, entry.mood, entry.activities_to_string()]
        entries_ws.append(row)
wb.save(OUTPUT_FILE)

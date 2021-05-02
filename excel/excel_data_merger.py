from openpyxl import Workbook, load_workbook
import sys

KEY_COLUMN_AMOUNT = 1
# KEEP_UNMATCHED_ROWS_FROM_FIRST = True

filename = sys.argv[1]

def getKeys(ws, row):
    keys = []
    for i in range(1, KEY_COLUMN_AMOUNT + 1):
        keys.append(ws.cell(row, i).value)
    return keys


def getRowWithKeys(ws, keys):
    for row in range(1, ws.max_row + 1):
        innerKeys = getKeys(ws, row)
        # Requires that the key columns are in the same order in both sheets
        if keys == innerKeys:
            return row
    return -1


wb = load_workbook(filename)

ws1 = wb["Sheet1"]
ws2 = wb["Sheet2"]
ws3 = wb.create_sheet("Result")

# Copy Headers
for column in range(1, ws2.max_column + 1):
    ws3.cell(1, column).value = ws2.cell(1, column).value


for row in range(2, ws1.max_row + 1):
    print("Row {} of {}".format(row, ws1.max_row))
    keys = getKeys(ws1, row)
    for i in range(0, KEY_COLUMN_AMOUNT):
        ws3.cell(row, i + 1).value = keys[i]

    # Copy data columns from sheet 1 to result sheet
    ws1DataColumnAmount = ws1.max_column - KEY_COLUMN_AMOUNT
    for column in range (KEY_COLUMN_AMOUNT + 1, ws1.max_column + 1):
        ws3.cell(row, column).value = ws1.cell(row, column).value

    innerRow = getRowWithKeys(ws2, keys)
    if innerRow == -1:
        continue
    for column in range(KEY_COLUMN_AMOUNT + 1, ws2.max_column + 1):
        ws3.cell(row, column + ws1DataColumnAmount).value = ws2.cell(innerRow, column).value


wb.save("result.xlsx")



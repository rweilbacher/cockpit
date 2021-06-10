import csv
from openpyxl import Workbook
import sys

source_file = sys.argv[1]
DESTINATION_FILE = source_file.replace(".csv", ".xlsx")
DELIMITER = "|"
CSV_DELIMITER = "^"
NEWLINE_REPLACEMENT = "\N{LATIN SMALL LETTER THORN}"


class ColumnInfo:
    def __init__(self, title, source_column, dest_column, append):
        self.title = title
        # Source column is not needed because it is the same as the index
        self.source_column = source_column
        self.dest_column = dest_column
        self.append = append

    def __str__(self):
        return "{title} - from: {source} to {dest} with append: {append}".format(title=self.title, source=self.source_column, dest=self.dest_column, append=self.append)


column_info = []
header_registry = {}


wb = Workbook()
ws = wb.active


# Sanitize CSV file
with open(source_file, "rb") as csv_file:
    csv_data = csv_file.read()

csv_data = csv_data.decode("utf-8")

new_csv_data = ""
column_amount = 0
column_counter = 0
for idx, char in enumerate(csv_data):
    print(idx)
    if char == "\r":
        # print("Skip \\r")
        pass
    elif char == CSV_DELIMITER:
        print("CSV Delimiter")
        new_csv_data += char
        column_counter += 1
    elif char == "\n":
        if column_amount == 0:
            # print("First column done")
            column_amount = column_counter
            new_csv_data += char
            column_counter = 0
        elif column_counter == column_amount:
            # print("Column done")
            new_csv_data += char
            column_counter = 0
        elif column_counter != column_amount:
            new_csv_data += NEWLINE_REPLACEMENT
            # print("Replaced newline")
    elif True:
        # I have no idea why else didn't work
        new_csv_data += char


temp_file = open("sanitized2.csv", "wb")
temp_file.write(new_csv_data.encode("utf-8"))
temp_file.close()

exit()

with open("sanitized.csv", "r", encoding="utf-8") as csv_file:
    # reader = csv.reader(csv_file, delimiter='^', doublequote='False', quotechar='', quoting=csv.QUOTE_NONE)
    reader = csv.reader(csv_file, delimiter=CSV_DELIMITER, quoting=csv.QUOTE_NONE, escapechar="\\")
    headers = next(reader)
    next_free_column = 0
    for i in range(0, len(headers)):
        header = headers[i]
        if header in header_registry:
            dest = header_registry[header]
            column_info.append(ColumnInfo(header, i, dest, True))
        else:
            header_registry[header] = next_free_column
            column_info.append(ColumnInfo(header, i, next_free_column, False))
            next_free_column += 1

    # Copy headers over
    for i, (key, value) in enumerate(header_registry.items()):
        ws.cell(1, i + 1).value = key

    # Copy values over
    for row_idx, row in enumerate(reader, start=1):
        print("Row: " + str(row_idx))
        for column_idx, value in enumerate(row):
            print("\tColumn: " + str(column_idx))
            info = column_info[column_idx]
            if info.append == True:
                # Only append the delimited value if it contains anything
                if str(value) != "":
                    cell = ws.cell(row_idx + 1, info.dest_column + 1)
                    new_value = str(cell.value) + DELIMITER + str(value)
                    cell.value = new_value
            else:
                ws.cell(row_idx + 1, info.dest_column + 1).value = value

print(header_registry)
for info in column_info:
    print(info)


wb.save(DESTINATION_FILE)

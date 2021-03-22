from openpyxl import Workbook, load_workbook

KEY_COLUMN_AMOUNT = 2
# KEEP_UNMATCHED_ROWS_FROM_FIRST = True

# TODO Verify data
# TODO Add column names from second sheet

def getKeys(ws, row):
    keys = []
    for i in range(1, KEY_COLUMN_AMOUNT + 1):
        keys.append(ws.cell(row, i).value)
    return keys


def getRowWithKeys(ws, keys):
    for row in range(1, ws.max_row):
        innerKeys = getKeys(ws, row)
        # Requires that the key columns are in the same order in both sheets
        if keys == innerKeys:
            return row
    return -1


wb = load_workbook("relevant_methods_2021_02.xlsx")

ws1 = wb["Sheet1"]
ws2 = wb["Sheet2"]
ws3 = wb.create_sheet("Result")

for row in range(1, ws1.max_row):
    keys = getKeys(ws1, row)
    for i in range(0, KEY_COLUMN_AMOUNT):
        ws3.cell(row, i + 1).value = keys[i]

    innerRow = getRowWithKeys(ws2, keys)
    if innerRow == -1:
        continue
    for column in range(KEY_COLUMN_AMOUNT + 1, ws2.max_column):
        ws3.cell(row, column).value = ws2.cell(innerRow, column).value


wb.save("result.xlsx")



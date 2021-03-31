import openpyxl


def isRowTrash(sheet, row):
    for header, column in headerColumns.items():
        cellValue = sheet.cell(row, column).value
        cellValue = cellValue.lower()
        for pattern in filter[header]:
            if pattern in cellValue:
                return True
    return False

filter = {"name": ["mitarbeiterlizenz", "use internal only", "gateway", "test"],
          "owner": ["support", "internal use only", "use internal only", "lucanet"],
          "status": ["locked", "activatable"],
          "type": ["employee_version", "test_version", "education_version"]}
headerColumns = {}

filename = "licenses.xlsx"


wb = openpyxl.load_workbook(filename)

sheet = wb.worksheets[0]
filteredSheet = wb.create_sheet(title="filtered")
trashSheet = wb.create_sheet(title="trash")

# Get the column index for all columns containing filters
for column in range(1, sheet.max_column + 1):
    value = sheet.cell(1, column).value
    if value in filter.keys():
        headerColumns[value] = column

trashRow = 1
filteredRow = 1
for row in range(1, sheet.max_row + 1):
    print("Row {} of {}".format(row, sheet.max_row))
    trash = isRowTrash(sheet, row)
    if trash is True:
        for column in range(1, sheet.max_column + 1):
            trashSheet.cell(trashRow, column).value = sheet.cell(row, column).value
        trashRow += 1
    else:
        for column in range(1, sheet.max_column + 1):
            filteredSheet.cell(filteredRow, column).value = sheet.cell(row, column).value
        filteredRow += 1

wb.save("filtered.xlsx")



